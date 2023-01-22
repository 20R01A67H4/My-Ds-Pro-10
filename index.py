from tkinter import *
import tkinter as tk
from tkinter.ttk import Combobox
import tkinter.messagebox as tmsg
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib


file=pathlib.Path('Backend.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Full Name"
    sheet['B1']="Roll Number"
    sheet['C1']="Department"
    sheet['D1']="Email Id"
    sheet['E1']="Skills"
    sheet['F1']="Certifications"
    sheet['G1']="Projects"
    sheet['H1']="CGPA"
    sheet['I1']="Backlogs"
    sheet['J1']="NumberOfBacklogs"
    sheet['K1']="PlacedInAnyCom"
    sheet['L1']="ComTotal"

    file.save('Backend.xlsx')

root=Tk()
root.title("Student Tracker Form")
root.geometry("1300x800")
root.minsize(1310,800)
root.maxsize(1310,800)
root.config(bg="#326273",pady=6)
icon=PhotoImage(file="icon.png")
root.iconphoto(False,icon)
root1=Frame(root).pack()
x=Label(root,text="Student Performance Tracker",font="arial 17",bg="red",fg="black",pady='10',width=1200).pack(side=TOP)

def submit():
    Name=NameValue.get()
    Roll=RollValue.get()
    Dept=DeptValue.get()
    Email=EmailValue.get()
    Skill=SkillsEntry.get(1.0,END)
    Certificates=CertificatesEntry.get(1.0,END)
    Projects=ProjectsEntry.get(1.0,END)
    CGPA=CGPAValue.get()
    Verify=BackVerify.get()
    Backlogs=NoOfBacklogs.get()
    Com=ComVerify.get()
    Companies=NoOfCompanies.get()


    if Name=='':
        tmsg.showerror("Error","Enter Your Name")
    elif Roll=='':
        tmsg.showerror("Error","Enter Your Roll Number")
    elif Dept=='':
        tmsg.showerror("Error","Enter Your Department")
    elif Email=='':
        tmsg.showerror("Error","Enter Your Email Id")
    elif Skill=='':
        tmsg.showerror("Error","Specify your skills")
    elif Certificates=='':
        tmsg.showerror("Error","Mention your certificates")
    elif Projects=='':
        tmsg.showerror("Error","Mention your projects")
    elif CGPA=='':
        tmsg.showerror("Error","Enter your current CGPA")
    elif Verify=='':
        tmsg.showerror("Error","Answer the questions'Do you have any Backlogs?'")
    elif Backlogs=='':
        tmsg.showerror("Error","Enter backlogs in total")
    elif Com=='':
        tmsg.showerror("Error","Answer the questions'Have you placed in any of the company?'")
    elif Companies=='':
        tmsg.showerror("Error","Specify number of companies you got placed")
    else:
        file=openpyxl.load_workbook('Backend.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=Name)
        sheet.cell(column=2,row=sheet.max_row,value=Roll)
        sheet.cell(column=3,row=sheet.max_row,value=Dept)
        sheet.cell(column=4,row=sheet.max_row,value=Email)
        sheet.cell(column=5,row=sheet.max_row,value=Skill)
        sheet.cell(column=6,row=sheet.max_row,value=Certificates)
        sheet.cell(column=7,row=sheet.max_row,value=Projects)
        sheet.cell(column=8,row=sheet.max_row,value=CGPA)
        sheet.cell(column=9,row=sheet.max_row,value=Verify)
        sheet.cell(column=10,row=sheet.max_row,value=Backlogs)
        sheet.cell(column=11,row=sheet.max_row,value=Com)
        sheet.cell(column=12,row=sheet.max_row,value=Companies)
        file.save('Backend.xlsx')
        tmsg.showinfo("Information","Details Successfully submitted")

def clear():
    NameEntry.delete(0,END)
    RollEntry.delete(0,END)
    DeptEntry.delete(0,END)
    EmailEntry.delete(0,END)
    CGPAEntry.delete(0,END)
    SkillsEntry.delete(1.0,END)
    CertificatesEntry.delete(1.0,END)
    ProjectsEntry.delete(1.0,END)
    
Label(root1,text="Full Name",font="20",bg="#326273",padx="70").place(x=50,y=100)
Label(root1,text="Roll Number",font="20",bg="#326273",padx="70").place(x=600,y=100)
Label(root1,text="Department",font="20",bg="#326273",padx="70").place(x=50,y=150)
Label(root1,text="Email Id",font="20",bg="#326273",padx="70").place(x=600,y=150)
Label(root1,text="Skills",font="20",bg="#326273",padx="70").place(x=50,y=200)
Label(root1,text="Certifications",font="20",bg="#326273",padx="70").place(x=600,y=200)
Label(root1,text="Projects",font="20",bg="#326273",padx="70").place(x=50,y=300)
Label(root1,text="Current CGPA",font="20",bg="#326273",padx="70").place(x=600,y=300)
Label(root1,text="Do you have any backlogs?",font="20",bg="#326273",padx="70").place(x=50,y=400)
Label(root1,text="Backlogs in total",font="20",bg="#326273",padx="70").place(x=600,y=400)
Label(root1,text="Have you placed in any of the company?",font="22220",bg="#326273",padx="70").place(x=50,y=450)
Label(root1,text="Number of companies got placed ",font="22220",bg="#326273",padx="70").place(x=600,y=450)
Button(root1,text="Submit",command=submit,font=("20"),bg="green",fg="white",padx="50").place(x=560,y=520)
Button(root1,text="Clear All",command=clear,font=("20"),bg="Blue",fg="white",padx="40").place(x=430,y=580)
Button(root1,text="Exit",command=lambda:root.destroy(),font=("20"),bg="red",fg="white",padx="50").place(x=700,y=580)
Label(root1,text="@tenkatirahulbabu                               ",font="arial 10",bg="red",fg="black",padx='0',width=1200,height=3).place(x=0,y=650,width=1400,height=30)

NameValue=StringVar()
RollValue=StringVar()
DeptValue=StringVar()
EmailValue=StringVar()
CGPAValue=StringVar()

NameEntry=Entry(root1,textvariable=NameValue,width=60,bd=3)
RollEntry=Entry(root1,textvariable=RollValue,width=60,bd=3)
DeptEntry=Entry(root1,textvariable=DeptValue,width=60,bd=3)
EmailEntry=Entry(root1,textvariable=EmailValue,width=60,bd=3)
SkillsEntry=Text(root1,width=45,height=4,bd=3)
CertificatesEntry=Text(root1,width=45,height=4,bd=3)
ProjectsEntry=Text(root1,width=45,height=4,bd=3)
CGPAEntry=Entry(root1,textvariable=CGPAValue,width=60,bd=3)

BackVerify=Combobox(root1,values=['No','Yes'],state="r",width=25)
BackVerify.place(x=430,y=400)
BackVerify.set('No')

NoOfBacklogs=Combobox(root1,values=['0','1','2','3','3+'],state="r",width=25)
NoOfBacklogs.place(x=930,y=400)
NoOfBacklogs.set('0')

ComVerify=Combobox(root1,values=['No','Yes'],state="r",width=25)
ComVerify.place(x=430,y=450)
ComVerify.set('No')

NoOfCompanies=Combobox(root1,values=['0','1','2','3','3+'],state="r",width=25)
NoOfCompanies.place(x=930,y=450)
NoOfCompanies.set('None')


NameEntry.place(x=240,y=100)
RollEntry.place(x=800,y=100)
DeptEntry.place(x=240,y=150)
EmailEntry.place(x=800,y=150)
SkillsEntry.place(x=240,y=200)
CertificatesEntry.place(x=800,y=200)
ProjectsEntry.place(x=240,y=300)
CGPAEntry.place(x=800,y=300)
root.mainloop()